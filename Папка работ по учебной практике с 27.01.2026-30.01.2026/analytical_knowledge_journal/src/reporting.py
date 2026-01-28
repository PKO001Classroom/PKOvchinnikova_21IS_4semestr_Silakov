# src/reporting.py
from datetime import datetime
import os
from pathlib import Path


class ReportGenerator:
    def __init__(self, db_manager, file_manager):
        self.db = db_manager
        self.fm = file_manager
        self.exports_dir = Path("exports")
        self.exports_dir.mkdir(exist_ok=True)

    def generate_excel_report(self):
        """Генерация Excel отчёта с графиками"""
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, Reference

            wb = Workbook()

            # Лист с данными
            ws_data = wb.active
            ws_data.title = "Данные"

            stats = self.db.get_total_stats()

            # Основные метрики
            ws_data.append(["Метрика", "Значение"])
            ws_data.append(["Всего конспектов", stats['total_notes']])
            ws_data.append(["Всего тегов", stats['total_tags']])
            ws_data.append(["Активность сегодня", stats['today_activity']])
            ws_data.append([])

            # Конспекты по категориям
            ws_data.append(["Конспекты по категориям"])
            ws_data.append(["Категория", "Количество"])
            for category, count in stats['notes_by_category'].items():
                ws_data.append([category, count])

            # Популярные теги
            ws_data.append([])
            ws_data.append(["Популярные теги"])
            ws_data.append(["Тег", "Использований"])
            for tag, count in stats['top_tags']:
                ws_data.append([tag, count])

            # Лист с графиками
            ws_charts = wb.create_sheet("Графики")

            # Данные для графика (активность по дням)
            ws_charts.append(["День", "Создано", "Обновлено", "Просмотрено"])

            # Получаем данные активности
            activity = self.db.get_activity_stats(7)

            # Если нет данных активности, создаём тестовые
            if activity and activity.get('daily_activity'):
                daily_data = activity['daily_activity']
                if daily_data:
                    for day_data in daily_data:
                        if isinstance(day_data, (list, tuple)) and len(day_data) >= 5:
                            ws_charts.append([
                                str(day_data[0]),  # Дата
                                day_data[2],  # CREATE
                                day_data[3],  # UPDATE
                                day_data[4]  # VIEW
                            ])
                else:
                    # Тестовые данные для демонстрации
                    ws_charts.append(["Пн", 2, 1, 5])
                    ws_charts.append(["Вт", 1, 3, 4])
                    ws_charts.append(["Ср", 3, 2, 6])
                    ws_charts.append(["Чт", 2, 1, 3])
                    ws_charts.append(["Пт", 4, 2, 7])
                    ws_charts.append(["Сб", 1, 0, 2])
                    ws_charts.append(["Вс", 0, 1, 1])
            else:
                # Тестовые данные
                ws_charts.append(["Пн", 2, 1, 5])
                ws_charts.append(["Вт", 1, 3, 4])
                ws_charts.append(["Ср", 3, 2, 6])
                ws_charts.append(["Чт", 2, 1, 3])
                ws_charts.append(["Пт", 4, 2, 7])
                ws_charts.append(["Сб", 1, 0, 2])
                ws_charts.append(["Вс", 0, 1, 1])

            # Создаем диаграмму
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Активность за неделю"
            chart.y_axis.title = 'Количество действий'
            chart.x_axis.title = 'Дни'

            # Определяем количество строк с данными
            max_row = ws_charts.max_row

            data = Reference(ws_charts, min_col=2, min_row=1, max_col=4, max_row=max_row)
            categories = Reference(ws_charts, min_col=1, min_row=2, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

            ws_charts.add_chart(chart, "F2")

            # Сохраняем
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = self.exports_dir / filename
            wb.save(filepath)

            print(f"✅ Excel отчёт с графиками создан: {filepath}")
            return str(filepath)

        except Exception as e:
            print(f"❌ Ошибка создания Excel отчёта: {e}")
            import traceback
            traceback.print_exc()
            return None

    def generate_pdf_report(self):
        """Генерация PDF отчёта"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            import io

            stats = self.db.get_total_stats()
            recent_notes = self.db.get_recent_notes(5)

            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = self.exports_dir / filename

            # Создаем PDF документ
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []

            # Заголовок
            title = Paragraph("Аналитический отчёт журнала знаний", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))

            # Дата генерации
            date_text = Paragraph(f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal'])
            story.append(date_text)
            story.append(Spacer(1, 24))

            # Основные метрики
            story.append(Paragraph("Основные метрики:", styles['Heading2']))
            story.append(Spacer(1, 12))

            metrics_data = [
                ["Показатель", "Значение"],
                ["Всего конспектов", str(stats['total_notes'])],
                ["Всего тегов", str(stats['total_tags'])],
                ["Активность сегодня", str(stats['today_activity'])]
            ]

            metrics_table = Table(metrics_data)
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            story.append(metrics_table)
            story.append(Spacer(1, 24))

            # Конспекты по категориям
            story.append(Paragraph("Конспекты по категориям:", styles['Heading2']))
            story.append(Spacer(1, 12))

            category_data = [["Категория", "Количество"]]
            for category, count in stats['notes_by_category'].items():
                category_data.append([category, str(count)])

            if len(category_data) > 1:  # Если есть данные кроме заголовка
                category_table = Table(category_data)
                category_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(category_table)
            else:
                story.append(Paragraph("Нет данных по категориям", styles['Normal']))

            story.append(Spacer(1, 24))

            # Популярные теги
            story.append(Paragraph("Популярные теги:", styles['Heading2']))
            story.append(Spacer(1, 12))

            tags_data = [["Тег", "Использований"]]
            for tag, count in stats['top_tags']:
                tags_data.append([tag, str(count)])

            if len(tags_data) > 1:
                tags_table = Table(tags_data)
                tags_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightyellow),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(tags_table)
            else:
                story.append(Paragraph("Теги ещё не добавлены", styles['Normal']))

            story.append(Spacer(1, 24))

            # Последние конспекты
            story.append(Paragraph("Последние конспекты:", styles['Heading2']))
            story.append(Spacer(1, 12))

            if recent_notes:
                notes_data = [["Название", "Категория", "Обновлён"]]
                for note in recent_notes:
                    notes_data.append([note['title'], note['category'], note['updated']])

                notes_table = Table(notes_data)
                notes_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightcoral),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightcyan),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke])
                ]))
                story.append(notes_table)
            else:
                story.append(Paragraph("Конспекты ещё не созданы", styles['Normal']))

            # Генерируем PDF
            doc.build(story)

            print(f"✅ PDF отчёт создан: {filepath}")
            return str(filepath)

        except ImportError as e:
            print(f"❌ Не установлена библиотека reportlab: {e}")
            print("Установите: pip install reportlab")
            return None
        except Exception as e:
            print(f"❌ Ошибка создания PDF отчёта: {e}")
            import traceback
            traceback.print_exc()
            return None


# Добавьте этот импорт в начало файла если SimpleDocTemplate не найден
from reportlab.platypus import SimpleDocTemplate